"""
Excel import script for Raid Shadow Legends Siege Assignment Web App.

Imports historical .xlsm siege files into the database.

Usage:
    python scripts/import_excel.py [<path>]
    python scripts/import_excel.py [<path>] --database-url postgresql+asyncpg://...

Where <path> is a single .xlsm file or a directory containing .xlsm files.
If omitted, falls back to the IMPORT_EXCEL_PATH environment variable.
"""

import argparse
import asyncio
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv

# Load .env from project root
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

# Add backend to path so we can import app models
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.db.base import Base  # noqa: F401 — needed to register all models with metadata
from app.db.seeds import seed_post_priority_config
from app.models.building import Building
from app.models.building_group import BuildingGroup
from app.models.enums import BuildingType, MemberRole, SiegeStatus
from app.models.member import Member
from app.models.position import Position
from app.models.post import Post
from app.models.post_condition import PostCondition
from app.models.post_priority_config import PostPriorityConfig
from app.models.siege import Siege
from app.models.siege_member import SiegeMember

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FILENAME_PATTERN = re.compile(r"clan_siege_(\d{2})_(\d{2})_(\d{4})")

BUILDING_ALIASES: dict[str, str] = {
    "Stronghold": "stronghold",
    "Mana Shrine": "mana_shrine",
    "Mana": "mana_shrine",
    "Magic Tower": "magic_tower",
    "Magic": "magic_tower",
    "Defense Tower": "defense_tower",
    "Defense": "defense_tower",
    "Post": "post",
}

ROLE_ALIASES: dict[str, str] = {
    "Heavy Hitter": "heavy_hitter",
    "Advanced": "advanced",
    "Medium": "medium",
    "Novice": "novice",
}

# building_type -> (base_group_count, last_group_slots)
BUILDING_TYPE_CONFIG: dict[str, tuple[int, int]] = {
    "stronghold": (4, 3),
    "mana_shrine": (2, 3),
    "magic_tower": (1, 2),
    "defense_tower": (1, 2),
    "post": (1, 1),
}

# building_type -> {total_position_count -> level}
BUILDING_LEVEL_MAP: dict[str, dict[int, int]] = {
    "stronghold": {12: 1, 16: 2, 18: 3, 22: 4, 25: 5, 30: 6},
    "mana_shrine": {6: 1, 7: 2, 9: 3, 11: 4, 13: 5, 15: 6},
    "magic_tower": {2: 1, 3: 2, 4: 3, 5: 4, 7: 5, 9: 6},
    "defense_tower": {2: 1, 3: 2, 4: 3, 6: 4, 9: 5, 12: 6},
    "post": {1: 1},
}

# post condition short keyword -> PostCondition.id  (case-insensitive lookup)
POST_CONDITION_KEYWORDS: dict[str, int] = {
    "telerian league": 1,
    "telerian": 1,
    "gaellen pact": 2,
    "the corrupted": 3,
    "corrupted": 3,
    "nyresan union": 4,
    "nyresan": 4,
    "hp": 5,
    "def": 6,
    "defense": 6,
    "support": 7,
    "atk": 8,
    "attack": 8,
    "banner lord": 9,
    "banner lords": 9,
    "bannerlord": 9,
    "bannerlords": 9,
    "high elves": 10,
    "sacred order": 11,
    "barbarian": 12,
    "barbarians": 12,
    "ogryn tribe": 13,
    "lizardmen": 14,
    "lizardman": 14,
    "skinwalker": 15,
    "skinwalkers": 15,
    "orc": 16,
    "orcs": 16,
    "tm reduce": 17,
    "tm-": 17,
    "tm fill": 18,
    "tm+": 18,
    "void": 19,
    "force": 20,
    "magic": 21,
    "spirit": 22,
    "demonspawn": 23,
    "undead horde": 24,
    "dark elves": 25,
    "knights revenant": 26,
    "knightrev": 26,
    "knightsrev": 26,
    "knight rev": 26,
    "cd increase": 27,
    "cd+": 27,
    "cd decrease": 28,
    "cd-": 28,
    "legendary": 29,
    "epic": 30,
    "epics only": 30,
    "epic's only": 30,
    "epics": 30,
    "rare": 31,
    "dwarves": 32,
    "dwarf": 32,
    "shadowkin": 33,
    "sylvan watcher": 34,
    "sylvan watchers": 34,
    "sheep": 35,
    "no revive": 36,
}

# ---------------------------------------------------------------------------
# Data classes for parsed sheet rows
# ---------------------------------------------------------------------------


@dataclass
class ParsedMember:
    name: str
    role: str
    power_level: Optional[str]
    discord_username: Optional[str]
    post_preference_keywords: list[str] = field(default_factory=list)


@dataclass
class ParsedAssignment:
    building_type: str  # canonical enum value e.g. "mana_shrine"
    building_number: int
    group_number: int
    position_number: int
    value: Optional[str]  # member name, "RESERVE", or None


@dataclass
class ParsedReserve:
    member_name: str
    attack_day: Optional[int]
    has_reserve_set: Optional[bool]


@dataclass
class ParsedPostConfig:
    post_number: int
    priority: int        # mapped to int: High=3, Medium=2, Low=1
    description: Optional[str]


@dataclass
class ParsedPostConditions:
    post_number: int          # 1-based, matches building_number for "post" buildings
    condition_keywords: list[str]   # up to 3, already split/stripped


@dataclass
class ImportStats:
    filename: str
    date: Optional[date] = None
    members_created: int = 0
    members_existing: int = 0
    buildings_created: int = 0
    positions_assigned: int = 0
    positions_reserve: int = 0
    positions_empty: int = 0
    siege_members_created: int = 0
    posts_created: int = 0
    preferences_set: int = 0
    post_configs_updated: int = 0
    post_conditions_set: int = 0
    siege_id: Optional[int] = None
    skipped: bool = False
    skip_reason: str = ""
    error: bool = False
    error_message: str = ""


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def parse_filename_date(filename: str) -> Optional[date]:
    """Extract the siege date from a filename like clan_siege_MM_DD_YYYY.xlsm."""
    match = FILENAME_PATTERN.search(filename)
    if not match:
        return None
    month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def map_role(display_role: str) -> Optional[str]:
    """Map a display role string to its enum value."""
    return ROLE_ALIASES.get(display_role)


def map_building_alias(raw_type: str) -> tuple[Optional[str], Optional[int]]:
    """
    Map a raw building type string to its canonical enum value and optional building number.

    Handles formats like "Magic Tower 1", "Post 12", "Mana Shrine 2" by stripping
    the trailing number. Returns (canonical_type, building_number) where building_number
    is extracted from the name if present, else None.
    """
    # Try exact match first
    if raw_type in BUILDING_ALIASES:
        return BUILDING_ALIASES[raw_type], None

    # Try stripping trailing number: "Magic Tower 1" -> "Magic Tower" + 1
    match = re.match(r"^(.+?)\s+(\d+)$", raw_type)
    if match:
        base_name = match.group(1)
        number = int(match.group(2))
        canonical = BUILDING_ALIASES.get(base_name)
        if canonical is not None:
            return canonical, number

    return None, None


def parse_members_sheet(ws) -> list[ParsedMember]:
    """
    Parse the 'Members' worksheet.

    Expected columns:
      A: name, B: level (ignored), C: player power, D: role, E: post restrictions
    Row 1 is header; data starts at row 2.
    """
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0] if len(row) > 0 else None
        if not name or not str(name).strip():
            continue
        if str(name).strip().upper() == "RESERVE":
            continue
        # Col B (index 1) is Level — skip it
        power_raw = row[2] if len(row) > 2 else None
        role_raw = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

        # Col E (index 4) is PostRestrictions — parse as comma-separated keywords
        post_restrictions_raw = row[4] if len(row) > 4 else None
        post_preference_keywords: list[str] = []
        if post_restrictions_raw is not None and str(post_restrictions_raw).strip():
            raw = str(post_restrictions_raw).replace(",", "/")
            tokens = raw.split("/")
            post_preference_keywords = [t.strip() for t in tokens if t.strip()]

        # Map text power ranges to power_level enum
        power_level: Optional[str] = None
        if power_raw is not None:
            power_str = str(power_raw).strip().lower()
            if "20" in power_str:
                power_level = "21_25m"
            elif "15" in power_str:
                power_level = "16_20m"
            elif "10" in power_str:
                power_level = "10_15m"
            elif "8" in power_str or "6" in power_str:
                power_level = "lt_10m"
            elif "<5" in power_str or "5.9" in power_str:
                power_level = "lt_10m"
            elif power_str in ("n/a", ""):
                power_level = None
            else:
                # Try numeric fallback
                try:
                    power_val = float(power_raw)
                    if power_val < 10_000_000:
                        power_level = "lt_10m"
                    elif power_val <= 15_000_000:
                        power_level = "10_15m"
                    elif power_val <= 20_000_000:
                        power_level = "16_20m"
                    elif power_val <= 25_000_000:
                        power_level = "21_25m"
                    else:
                        power_level = "gt_25m"
                except (TypeError, ValueError):
                    power_level = None

        members.append(
            ParsedMember(
                name=str(name).strip(),
                role=role_raw,
                power_level=power_level,
                discord_username=None,
                post_preference_keywords=post_preference_keywords,
            )
        )
    return members


def parse_assignments_sheet(ws) -> list[ParsedAssignment]:
    """
    Parse the 'Assignments' worksheet in its visual grid format.

    Layout:
      Row 1: header row (Location | Group | Assigned | None | None | ...)
      Row 2: position sub-headers (None | None | 1 | 2 | 3 | ...)
      Row 3+: data rows
        Col A (index 0): building type display name — only filled on the FIRST row
                         of each new building instance; blank for subsequent groups
                         of the same building.
        Col B (index 1): group number (integer)
        Col C+ (index 2+): member name (or None) at position 1, 2, 3, ...

    Building number tracking:
      Each time col A is non-blank for the same canonical building type, it signals
      a new building instance and the building_number increments for that type.
      When col A is blank, the previous building type and number are carried forward.
    """
    assignments = []

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(all_rows) < 3:
        return assignments

    # Row 2 (index 1): position sub-headers starting at col C (index 2).
    # Read how many position columns the sheet actually has.
    subheader_row = all_rows[1]
    position_count = 0
    for col_idx in range(2, len(subheader_row)):
        cell = subheader_row[col_idx]
        if cell is not None:
            position_count += 1
        else:
            # Stop at first blank — positions are contiguous
            break
    # Fallback: if sub-header row is entirely blank in cols C+, infer from data width
    if position_count == 0:
        for data_row in all_rows[2:]:
            if len(data_row) > 2:
                position_count = max(position_count, len(data_row) - 2)

    # Track current building context as we scan data rows
    current_type: Optional[str] = None   # canonical building type
    # Map canonical type -> how many building instances of that type we've seen so far
    building_number_counter: dict[str, int] = {}
    current_building_number: int = 0

    for row in all_rows[2:]:  # data starts at row 3 (index 2)
        # Col A: building type name (may be blank)
        raw_type_cell = row[0] if len(row) > 0 else None
        raw_type = str(raw_type_cell).strip() if raw_type_cell is not None else ""

        # Col B: group number (may be "N/A" for single-group buildings)
        group_raw = row[1] if len(row) > 1 else None
        if group_raw is None:
            continue
        try:
            group_number = int(group_raw)
        except (TypeError, ValueError):
            # "N/A" or similar — treat as group 1
            group_number = 1

        if raw_type:
            # Non-blank col A: new building instance
            canonical_type, explicit_number = map_building_alias(raw_type)
            if canonical_type is None:
                # Unknown building type — skip this row and clear context so
                # subsequent blank-A rows don't inherit a bad type
                current_type = None
                current_building_number = 0
                continue
            current_type = canonical_type
            if explicit_number is not None:
                # Name had a number like "Magic Tower 1" — use it directly
                current_building_number = explicit_number
            else:
                # No number in name — auto-increment
                building_number_counter[canonical_type] = building_number_counter.get(canonical_type, 0) + 1
                current_building_number = building_number_counter[canonical_type]
        else:
            # Blank col A: continuing the same building as the last non-blank row
            if current_type is None:
                # No building context yet — skip
                continue

        # Emit one ParsedAssignment per position column
        for pos_idx in range(position_count):
            col_idx = 2 + pos_idx
            value_raw = row[col_idx] if col_idx < len(row) else None

            value: Optional[str] = None
            if value_raw is not None and str(value_raw).strip():
                value = str(value_raw).strip()

            assignments.append(
                ParsedAssignment(
                    building_type=current_type,
                    building_number=current_building_number,
                    group_number=group_number,
                    position_number=pos_idx + 1,
                    value=value,
                )
            )

    return assignments


def parse_reserves_sheet(ws) -> list[ParsedReserve]:
    """
    Parse the 'Reserves' worksheet.

    Expected columns:
      A: member_name, B: attack_day, C: has_reserve_set
    Row 1 is header; data starts at row 2.
    """
    reserves = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_raw = row[0] if len(row) > 0 else None
        if not name_raw or not str(name_raw).strip():
            continue

        attack_day_raw = row[1] if len(row) > 1 else None
        has_reserve_raw = row[2] if len(row) > 2 else None

        attack_day: Optional[int] = None
        if attack_day_raw is not None:
            try:
                val = int(attack_day_raw)
                if val in (1, 2):
                    attack_day = val
            except (TypeError, ValueError):
                pass

        has_reserve_set: Optional[bool] = None
        if has_reserve_raw is not None:
            normalized = str(has_reserve_raw).strip().lower()
            if normalized == "yes":
                has_reserve_set = True
            elif normalized == "no":
                has_reserve_set = False

        reserves.append(
            ParsedReserve(
                member_name=str(name_raw).strip(),
                attack_day=attack_day,
                has_reserve_set=has_reserve_set,
            )
        )
    return reserves


def parse_posts_sheet_config(ws) -> list[ParsedPostConfig]:
    """
    Parse post priority and descriptions from the Posts sheet, rows 2–17, columns B–D.

    Layout is hierarchical: a row whose col B is a non-numeric string is a priority
    section header ("High Priority", "Medium Priority", "Low Priority"). A row whose
    col B is an integer is a post data row with col C as the description.

    Priority integer mapping: High=3, Medium=2, Low=1. Default priority is Low (1).
    """
    _PRIORITY_MAP = {"high": 3, "medium": 2, "low": 1}
    configs: list[ParsedPostConfig] = []
    current_priority = 1  # default: Low

    for row in ws.iter_rows(min_row=2, max_row=17, min_col=2, max_col=4, values_only=True):
        cell_b = row[0]
        cell_c = row[1]

        if cell_b is None:
            continue

        # Try to interpret as an integer (post number)
        try:
            post_number = int(cell_b)
        except (TypeError, ValueError):
            # Non-numeric string — treat as a priority section header
            label = str(cell_b).lower()
            for keyword, pval in _PRIORITY_MAP.items():
                if keyword in label:
                    current_priority = pval
                    break
            continue

        description = str(cell_c).strip() if cell_c is not None else None
        if description == "":
            description = None
        configs.append(ParsedPostConfig(
            post_number=post_number,
            priority=current_priority,
            description=description,
        ))

    return configs


def parse_posts_sheet_conditions(ws) -> list[ParsedPostConditions]:
    """
    Parse post active conditions from the Posts sheet, rows 34–51, columns D–F.

    Row 34 = Post 1, row 35 = Post 2, …, row 51 = Post 18. Each row has up to 3
    condition keywords in columns D, E, F. Each cell is a single keyword (not
    comma/slash-separated). Only posts with at least one non-empty keyword are included.
    """
    results: list[ParsedPostConditions] = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=34, max_row=51, min_col=4, max_col=6, values_only=True),
        start=34,
    ):
        post_number = row_index - 33  # row 34 -> post 1
        condition_keywords: list[str] = []
        for cell in row:
            if cell is not None:
                stripped = str(cell).strip()
                if stripped:
                    condition_keywords.append(stripped)

        if condition_keywords:
            results.append(ParsedPostConditions(
                post_number=post_number,
                condition_keywords=condition_keywords,
            ))

    return results


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


def build_group_structure(building_type: str) -> list[int]:
    """
    Return a list of slot_counts for each group of the given building type.

    e.g. stronghold -> [3, 3, 3, 3]  (4 groups, last group also has 3)
    e.g. magic_tower -> [3, 2]        (but only 1 group for magic_tower? see config)

    Actually per config:
      stronghold:   base_groups=4, last_slots=3  -> [3, 3, 3, 3]
      mana_shrine:  base_groups=2, last_slots=3  -> [3, 3]
      magic_tower:  base_groups=1, last_slots=2  -> [2]
      defense_tower:base_groups=1, last_slots=2  -> [2]
      post:         base_groups=1, last_slots=1  -> [1]

    All groups except the last have slot_count=3. The last group has last_slots.
    """
    base_groups, last_slots = BUILDING_TYPE_CONFIG[building_type]
    if base_groups == 1:
        return [last_slots]
    return [3] * (base_groups - 1) + [last_slots]


def compute_building_group_structure(
    assignments: list[ParsedAssignment],
    building_type: str,
    building_number: int,
) -> dict[int, int]:
    """
    Scan the given list of ParsedAssignments and return {group_number: slot_count}
    for the specified building instance (identified by building_type + building_number).

    The Assignments sheet has a single global position_count (number of position columns),
    which is the maximum across all building types — typically 3.  Buildings with fewer
    real slots per group (Magic Tower: 2, Post: 1) receive trailing all-None position
    columns that must not be counted as real slots.

    Resolution strategy:
      - Non-last groups always have exactly 3 slots.
      - For the last group, start from base_last_slots as the initial minimum.
      - Walk the last group's slot count DOWN (3 → 2 → 1) until the resulting total
        maps to a recognised level in BUILDING_LEVEL_MAP.  This handles buildings at
        upgrade levels where the last group has fewer slots than base_last_slots (e.g.
        Stronghold level 2 has 16 teams = 5 full groups + 1 slot in the last group).
      - If the filled-position evidence pushes above the initial minimum (e.g. a member
        is assigned to position 3 of a magic_tower group), we use max_filled instead
        since the level map will accept that higher total.
      - Post (base_last=1): always stays at 1 — no trailing positions are ever filled.

    This means:
      - Magic Tower (base_last=2): starts at 2; rises to 3 only if position 3 is filled;
        may stay at 1 when the last group has only 1 real slot (level 3: 4 = 1×3 + 1×1).
      - Stronghold / Mana Shrine (base_last=3): starts at 3 but may drop to 2 or 1 when
        the building is at a level where the last group is smaller.
      - Post (base_last=1): stays at 1 always.
    """
    building_assignments = [
        a for a in assignments
        if a.building_type == building_type and a.building_number == building_number
    ]
    if not building_assignments:
        return {}

    _, base_last_slots = BUILDING_TYPE_CONFIG.get(building_type, (1, 3))

    # Ordered group numbers so we can identify the last group.
    group_numbers = sorted(set(a.group_number for a in building_assignments))
    n_groups = len(group_numbers)

    # For each group, find the highest position that has any non-None value.
    max_filled: dict[int, int] = {}
    for a in building_assignments:
        if a.value is not None and a.position_number > max_filled.get(a.group_number, 0):
            max_filled[a.group_number] = a.position_number

    # Build the initial group_structure using base_last_slots as the floor for the
    # last group (and 3 for all non-last groups).
    group_structure: dict[int, int] = {}
    last_gnum: Optional[int] = None
    for i, gnum in enumerate(group_numbers):
        is_last = (i == n_groups - 1)
        base_slots = base_last_slots if is_last else 3
        group_structure[gnum] = max(base_slots, max_filled.get(gnum, 0))
        if is_last:
            last_gnum = gnum

    # The last group's slot count was floored to base_last_slots, but at upgrade levels
    # the real last-group slot count can be lower.  Find the smallest slot count for the
    # last group that is (a) >= the highest filled position and (b) produces a total
    # that maps to a recognised level.  We prefer the smallest valid count so we do not
    # accidentally assign a higher level than the building actually is.
    level_map = BUILDING_LEVEL_MAP.get(building_type, {})
    if last_gnum is not None and level_map:
        # The evidence floor: we cannot go below the highest filled position.
        evidence_floor = max(1, max_filled.get(last_gnum, 1))
        # Try each candidate from the evidence floor up to the current (base-floored) value.
        current_last_slots = group_structure[last_gnum]
        best_last_slots = current_last_slots  # fallback: keep the base-floored value
        for candidate in range(evidence_floor, current_last_slots + 1):
            group_structure[last_gnum] = candidate
            if sum(group_structure.values()) in level_map:
                best_last_slots = candidate
                break
        group_structure[last_gnum] = best_last_slots

    return group_structure


def infer_building_level(building_type: str, group_structure: dict[int, int]) -> int:
    """
    Sum all slot counts in group_structure and look up the level for that building type.
    Falls back to level 1 if the total is not in the level map.
    """
    total_positions = sum(group_structure.values())
    level_map = BUILDING_LEVEL_MAP.get(building_type, {})
    return level_map.get(total_positions, 1)


async def get_or_create_member(
    session: AsyncSession,
    parsed: ParsedMember,
) -> tuple[Member, bool]:
    """
    Return (member, created).
    Looks up by name (case-insensitive). Creates if not found.
    """
    result = await session.execute(
        select(Member).where(Member.name.ilike(parsed.name))
    )
    existing = result.scalar_one_or_none()
    if existing:
        return existing, False

    role_value = map_role(parsed.role)
    if role_value is None:
        # Default to novice if role is unrecognised
        role_value = "novice"

    member = Member(
        name=parsed.name,
        role=MemberRole(role_value),
        power_level=parsed.power_level,
        discord_username=parsed.discord_username,
        is_active=True,
    )
    session.add(member)
    await session.flush()  # get the id without committing
    return member, True


async def create_building_with_groups_and_positions(
    session: AsyncSession,
    siege_id: int,
    building_type: str,
    building_number: int,
    level: int,
    group_structure: dict[int, int],
) -> tuple[Building, dict[tuple[int, int], Position]]:
    """
    Create a Building, its BuildingGroups, and their Positions.

    Uses the provided level and group_structure ({group_number: slot_count}) instead
    of the hardcoded BUILDING_TYPE_CONFIG so that higher-level buildings are created
    with the correct number of groups and slots as observed in the sheet.

    Returns (building, positions_map) where positions_map is
    {(group_number, position_number): Position}.
    """
    building = Building(
        siege_id=siege_id,
        building_type=BuildingType(building_type),
        building_number=building_number,
        level=level,
        is_broken=False,
    )
    session.add(building)
    await session.flush()

    positions_map: dict[tuple[int, int], Position] = {}

    for group_num in sorted(group_structure.keys()):
        slot_count = group_structure[group_num]
        group = BuildingGroup(
            building_id=building.id,
            group_number=group_num,
            slot_count=slot_count,
        )
        session.add(group)
        await session.flush()

        for pos_num in range(1, slot_count + 1):
            position = Position(
                building_group_id=group.id,
                position_number=pos_num,
            )
            session.add(position)
            await session.flush()
            positions_map[(group_num, pos_num)] = position

    return building, positions_map


# ---------------------------------------------------------------------------
# Core import logic
# ---------------------------------------------------------------------------


async def import_file(
    session: AsyncSession,
    filepath: Path,
    is_most_recent: bool = False,
) -> ImportStats:
    """Import a single .xlsm file. Returns ImportStats."""
    import openpyxl

    stats = ImportStats(filename=filepath.name)

    # 1. Extract date from filename
    siege_date = parse_filename_date(filepath.name)
    if siege_date is None:
        stats.skipped = True
        stats.skip_reason = f"Filename does not match pattern clan_siege_DD_MM_YYYY: {filepath.name}"
        return stats

    stats.date = siege_date

    # Load workbook
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        stats.error = True
        stats.error_message = f"Failed to open workbook: {exc}"
        return stats

    # 2. Parse sheets
    try:
        members_ws = wb["Members"]
        assignments_ws = wb["Assignments"]
        reserves_ws = wb["Reserves"]
    except KeyError as exc:
        stats.error = True
        stats.error_message = f"Missing required sheet: {exc}. Available: {wb.sheetnames}"
        wb.close()
        return stats

    parsed_members = parse_members_sheet(members_ws)
    parsed_assignments = parse_assignments_sheet(assignments_ws)
    parsed_reserves = parse_reserves_sheet(reserves_ws)

    # Parse Posts sheet (optional — silently skip if absent)
    parsed_post_configs: list[ParsedPostConfig] = []
    parsed_post_conditions: list[ParsedPostConditions] = []
    try:
        posts_ws = wb["Posts"]
        parsed_post_configs = parse_posts_sheet_config(posts_ws)
        parsed_post_conditions = parse_posts_sheet_conditions(posts_ws)
    except KeyError:
        pass

    wb.close()

    print(f"  Parsed: {len(parsed_members)} members, {len(parsed_assignments)} assignments, {len(parsed_reserves)} reserves")

    # 3. Upsert members
    member_name_map: dict[str, Member] = {}  # lowercase name -> Member
    for pm in parsed_members:
        member, created = await get_or_create_member(session, pm)
        member_name_map[member.name.lower()] = member
        if created:
            stats.members_created += 1
        else:
            stats.members_existing += 1
            if is_most_recent:
                # Update role and power_level from most recent data
                role_value = map_role(pm.role)
                if role_value:
                    member.role = MemberRole(role_value)
                if pm.power_level:
                    member.power_level = pm.power_level

    # Load PostCondition rows — needed for member preferences (most_recent) and for
    # post active conditions (every file).  Fetch once and reuse.
    pc_result = await session.execute(select(PostCondition))
    post_conditions: dict[int, PostCondition] = {
        pc.id: pc for pc in pc_result.scalars().all()
    }

    # 3b. Import post preferences (most recent file only)
    if is_most_recent:

        # Build a normalised keyword -> condition_id map
        kw_lower: dict[str, int] = {k.lower(): v for k, v in POST_CONDITION_KEYWORDS.items()}

        for pm in parsed_members:
            if not pm.post_preference_keywords:
                continue
            member = member_name_map.get(pm.name.lower())
            if member is None:
                continue

            # Resolve keywords to condition IDs
            condition_ids: list[int] = []
            for kw in pm.post_preference_keywords:
                cid = kw_lower.get(kw.lower())
                if cid is None:
                    print(f"  Warning: unknown post preference keyword '{kw}' for member '{pm.name}' — skipping")
                    continue
                if cid in post_conditions:
                    condition_ids.append(cid)
                else:
                    print(f"  Warning: PostCondition id={cid} not found in DB for keyword '{kw}' — skipping")

            # Delete existing preferences for this member
            await session.execute(
                text("DELETE FROM member_post_preference WHERE member_id = :mid"),
                {"mid": member.id},
            )

            # Insert new preferences
            inserted = 0
            for cid in condition_ids:
                await session.execute(
                    text(
                        "INSERT INTO member_post_preference (member_id, post_condition_id) "
                        "VALUES (:mid, :cid) ON CONFLICT DO NOTHING"
                    ),
                    {"mid": member.id, "cid": cid},
                )
                inserted += 1
            stats.preferences_set += inserted

        # 3c. Update PostPriorityConfig priority/description (most recent file only)
        if is_most_recent and parsed_post_configs:
            ppc_result = await session.execute(select(PostPriorityConfig))
            ppc_map: dict[int, PostPriorityConfig] = {
                ppc.post_number: ppc for ppc in ppc_result.scalars().all()
            }
            for pc in parsed_post_configs:
                existing_ppc = ppc_map.get(pc.post_number)
                if existing_ppc is not None:
                    existing_ppc.priority = pc.priority
                    existing_ppc.description = pc.description
                    stats.post_configs_updated += 1

    # 4. Create Siege
    siege = Siege(
        date=siege_date,
        status=SiegeStatus.complete,
        defense_scroll_count=3,
    )
    session.add(siege)
    await session.flush()
    stats.siege_id = siege.id

    # 5. Create Buildings, groups, and positions
    # Collect unique (building_type, building_number) pairs in insertion order
    seen_buildings: dict[tuple[str, int], tuple[Building, dict[tuple[int, int], Position]]] = {}
    for assignment in parsed_assignments:
        key = (assignment.building_type, assignment.building_number)
        if key not in seen_buildings:
            group_structure = compute_building_group_structure(
                parsed_assignments,
                assignment.building_type,
                assignment.building_number,
            )
            level = infer_building_level(assignment.building_type, group_structure)
            building, positions_map = await create_building_with_groups_and_positions(
                session,
                siege.id,
                assignment.building_type,
                assignment.building_number,
                level=level,
                group_structure=group_structure,
            )
            seen_buildings[key] = (building, positions_map)
            stats.buildings_created += 1

    # 6. Assign positions
    for assignment in parsed_assignments:
        key = (assignment.building_type, assignment.building_number)
        if key not in seen_buildings:
            continue
        _, positions_map = seen_buildings[key]
        pos_key = (assignment.group_number, assignment.position_number)
        position = positions_map.get(pos_key)
        if position is None:
            # Position doesn't exist in our structure (data mismatch), skip
            continue

        if assignment.value is None:
            stats.positions_empty += 1
        elif assignment.value.upper() == "RESERVE":
            position.is_reserve = True
            stats.positions_reserve += 1
        else:
            # Member assignment — look up by name
            member = member_name_map.get(assignment.value.lower())
            if member is not None:
                position.member_id = member.id
                stats.positions_assigned += 1
            else:
                # Member appeared in assignments but not in Members sheet
                # Try a DB lookup as fallback
                result = await session.execute(
                    select(Member).where(Member.name.ilike(assignment.value))
                )
                fallback_member = result.scalar_one_or_none()
                if fallback_member:
                    position.member_id = fallback_member.id
                    member_name_map[fallback_member.name.lower()] = fallback_member
                    stats.positions_assigned += 1
                else:
                    stats.positions_empty += 1

    # 7. Create SiegeMember rows
    for reserve in parsed_reserves:
        member = member_name_map.get(reserve.member_name.lower())
        if member is None:
            # Try DB lookup
            result = await session.execute(
                select(Member).where(Member.name.ilike(reserve.member_name))
            )
            member = result.scalar_one_or_none()
            if member:
                member_name_map[member.name.lower()] = member

        if member is None:
            continue

        siege_member = SiegeMember(
            siege_id=siege.id,
            member_id=member.id,
            attack_day=reserve.attack_day,
            has_reserve_set=reserve.has_reserve_set,
        )
        session.add(siege_member)
        stats.siege_members_created += 1

    # 8. Create Post rows for all post-type buildings
    # Build a lookup from post_number -> ParsedPostConditions for condition insertion
    post_conditions_lookup: dict[int, ParsedPostConditions] = {
        ppc.post_number: ppc for ppc in parsed_post_conditions
    }
    # Full description -> condition_id lookup (post conditions use full description strings).
    desc_to_condition_id: dict[str, int] = {
        pc.description.strip().lower(): pc.id for pc in post_conditions.values()
    }

    for (building_type, building_number), (building, _) in seen_buildings.items():
        if building_type == "post":
            post = Post(
                siege_id=siege.id,
                building_id=building.id,
                priority=0,
                description=None,
            )
            session.add(post)
            await session.flush()
            stats.posts_created += 1

            # Insert active conditions for this post if available.
            # Cell values are full PostCondition description strings.
            ppc_entry = post_conditions_lookup.get(building_number)
            if ppc_entry:
                for kw in ppc_entry.condition_keywords:
                    cid = desc_to_condition_id.get(kw.strip().lower())
                    if cid is None:
                        print(f"  Warning: unknown post condition '{kw}' for post {building_number} — skipping")
                        continue
                    await session.execute(
                        text(
                            "INSERT INTO post_active_condition (post_id, post_condition_id) "
                            "VALUES (:pid, :cid) ON CONFLICT DO NOTHING"
                        ),
                        {"pid": post.id, "cid": cid},
                    )
                    stats.post_conditions_set += 1

    # 9. Mark members not in most recent siege as inactive
    if is_most_recent:
        all_members_result = await session.execute(
            select(Member).where(Member.is_active == True)  # noqa: E712
        )
        all_active_members = all_members_result.scalars().all()
        deactivated = 0
        for m in all_active_members:
            if m.name.lower() not in member_name_map:
                m.is_active = False
                deactivated += 1
        if deactivated:
            print(f"  Deactivated {deactivated} members not in most recent siege")

    await session.flush()

    if is_most_recent:
        print(f"  (Most recent file — updated existing member roles/power levels, deactivated absent members)")

    return stats


async def import_files(
    database_url: str,
    filepaths: list[Path],
) -> None:
    """Import a list of .xlsm files into the database."""
    engine = create_async_engine(database_url, echo=False)
    SessionFactory = async_sessionmaker(engine, expire_on_commit=False)

    total_imported = 0
    total_skipped = 0
    total_errors = 0

    # Ensure PostPriorityConfig rows exist before processing any file.
    # seed_post_priority_config uses ON CONFLICT DO NOTHING so it is idempotent.
    async with SessionFactory() as session:
        async with session.begin():
            await seed_post_priority_config(session)

    for i, filepath in enumerate(filepaths):
        is_last = (i == len(filepaths) - 1)
        print(f"Importing: {filepath.name}")

        async with SessionFactory() as session:
            async with session.begin():
                try:
                    stats = await import_file(session, filepath, is_most_recent=is_last)
                except Exception as exc:
                    await session.rollback()
                    print(f"  ERROR: Unexpected error — {exc}")
                    total_errors += 1
                    continue

        if stats.skipped:
            print(f"  SKIPPED: {stats.skip_reason}")
            total_skipped += 1
            continue

        if stats.error:
            print(f"  ERROR: {stats.error_message}")
            total_errors += 1
            continue

        print(f"  Date: {stats.date}")
        print(
            f"  Members: {stats.members_created} created, "
            f"{stats.members_existing} existing"
        )
        print(f"  Buildings: {stats.buildings_created} created")
        print(
            f"  Positions: {stats.positions_assigned} assigned, "
            f"{stats.positions_reserve} RESERVE, "
            f"{stats.positions_empty} empty"
        )
        print(f"  Siege members: {stats.siege_members_created} created")
        print(f"  Posts: {stats.posts_created} created")
        print(f"  Post conditions: {stats.post_conditions_set} active conditions imported")
        if is_last:
            print(f"  Post preferences: {stats.preferences_set} preference records set")
            print(f"  Post configs: {stats.post_configs_updated} priority/description records updated")
        print(f"  + Siege #{stats.siege_id} created (status: complete)")
        total_imported += 1

    print()
    print(
        f"Summary: {total_imported} files imported, "
        f"{total_skipped} skipped, "
        f"{total_errors} errors"
    )

    await engine.dispose()


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def collect_xlsm_files(path: Path) -> list[Path]:
    """Return a sorted list of .xlsm files from a file path or directory."""
    if path.is_file():
        if path.suffix.lower() == ".xlsm":
            return [path]
        print(f"Warning: {path} is not a .xlsm file, skipping.")
        return []
    if path.is_dir():
        files = sorted(path.glob("*.xlsm"))
        if not files:
            print(f"Warning: no .xlsm files found in {path}")
        return list(files)
    print(f"Error: {path} does not exist.")
    sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import historical .xlsm siege files into the database."
    )
    parser.add_argument(
        "path",
        type=Path,
        nargs="?",
        default=None,
        help="Path to a single .xlsm file or a directory containing .xlsm files. "
        "Falls back to IMPORT_EXCEL_PATH env var if not provided.",
    )
    parser.add_argument(
        "--database-url",
        default=os.environ.get("IMPORT_DATABASE_URL", os.environ.get("DATABASE_URL")),
        help="Database connection URL (defaults to IMPORT_DATABASE_URL, then DATABASE_URL env var).",
    )
    args = parser.parse_args()

    if not args.database_url:
        print(
            "Error: DATABASE_URL is not set. "
            "Provide it via --database-url or the DATABASE_URL environment variable."
        )
        sys.exit(1)

    path = args.path or os.environ.get("IMPORT_EXCEL_PATH")
    if not path:
        print(
            "Error: No import path provided. "
            "Pass it as an argument or set the IMPORT_EXCEL_PATH environment variable."
        )
        sys.exit(1)

    filepaths = collect_xlsm_files(Path(path))
    if not filepaths:
        sys.exit(0)

    asyncio.run(import_files(args.database_url, filepaths))


if __name__ == "__main__":
    main()
